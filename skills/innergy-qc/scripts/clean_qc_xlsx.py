#!/usr/bin/env python3
"""
clean_qc_xlsx.py
Removes empty columns from QC Comparison sheet and colors cells.
Produces clean side-by-side: INNERGY X,Y,Z,Qty | OpenClaw X,Y,Z,Qty
Green = match, Red = mismatch, Grey = no OC data.

Usage:
    python3 scripts/clean_qc_xlsx.py \
        --xlsx path/to/innergy_qc_comparison_raw.xlsx \
        --output path/to/innergy_qc_comparison.xlsx
"""

import argparse
import openpyxl
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side

GREEN_FILL  = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
RED_FILL    = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
ORANGE_FILL = PatternFill(start_color="FFEB9C", end_color="FFEB9C", fill_type="solid")
GREY_FILL   = PatternFill(start_color="E8EDF2", end_color="E8EDF2", fill_type="solid")
HEADER_FILL = PatternFill(start_color="1E3A5F", end_color="1E3A5F", fill_type="solid")
thin = Side(style="thin", color="BBBBBB")
border = Border(left=thin, right=thin, top=thin, bottom=thin)


def num(v):
    if v is None:
        return None
    if isinstance(v, (int, float)):
        return float(v)
    try:
        return float(str(v).strip())
    except (ValueError, TypeError):
        return None


def color_cell(cell, ig_val, oc_val, tol=0.1):
    cell.border = border
    ig_n = num(ig_val)
    oc_n = num(oc_val)
    if oc_n is None:
        cell.fill = GREY_FILL
        return
    if ig_n is None or oc_n is None:
        cell.fill = RED_FILL
        return
    if abs(ig_n - oc_n) <= tol:
        cell.fill = GREEN_FILL
    else:
        cell.fill = RED_FILL


def main(xlsx_path, output_path):
    wb = openpyxl.load_workbook(xlsx_path)
    ws = wb["QC Comparison"]

    # Delete trailing empty columns (13+)
    for col in range(ws.max_column, 12, -1):
        ws.delete_cols(col)

    # Delete empty columns among 9-12 (OC X, OC Y, OC Z, OC Qty)
    # Check each and delete if completely empty
    for col in [9, 10, 12]:  # OC X, OC Y, OC Qty — may be empty
        is_empty = all(ws.cell(row=r, column=col).value is None for r in range(2, ws.max_row + 1))
        if is_empty:
            ws.delete_cols(col)

    # Now layout should be: 1-8 (INNERGY) + 9=OC Z + 10=Status + 11=Notes
    # Rename remaining OC column headers
    ws.cell(row=1, column=9).value = "OpenClaw Z"
    ws.cell(row=1, column=9).border = border
    ws.cell(row=1, column=9).alignment = Alignment(horizontal="center")
    ws.cell(row=1, column=9).font = Font(bold=True, color="FFFFFF")

    # Delete Status(10) and Notes(11) columns if present
    for col in [11, 10]:
        is_empty = all(ws.cell(row=r, column=col).value is None for r in range(1, ws.max_row + 1))
        if is_empty:
            ws.delete_cols(col)

    # Final column headers
    headers = ["Line #", "Name", "Location", "Origin", "X", "Y", "Z", "Qty", "OpenClaw Z"]
    for col_idx, hdr in enumerate(headers, 1):
        c = ws.cell(row=1, column=col_idx)
        c.value = hdr
        c.fill = HEADER_FILL
        c.font = Font(bold=True, color="FFFFFF")
        c.border = border
        c.alignment = Alignment(horizontal="center")

    # Color each OC cell
    for row_idx in range(2, ws.max_row + 1):
        z_ig = ws.cell(row=row_idx, column=7).value
        z_oc = ws.cell(row=row_idx, column=9).value
        cell = ws.cell(row=row_idx, column=9)
        if z_oc is not None:
            if num(z_ig) == num(z_oc):
                cell.fill = GREEN_FILL
            else:
                cell.fill = RED_FILL
        else:
            cell.fill = GREY_FILL
        cell.border = border

    # Style header row
    for col in range(1, ws.max_column + 1):
        c = ws.cell(row=1, column=col)
        c.fill = HEADER_FILL
        c.font = Font(bold=True, color="FFFFFF")
        c.border = border

    # Add legend
    ws_sum = wb["QC Summary"]
    ws_sum["A1"] = "QC Summary — 595 Market Street Run 3"
    ws_sum["A1"].font = Font(bold=True, size=14)
    ws_sum["A2"] = "Legend:"
    ws_sum["A3"] = "Green = OpenClaw Z matches INNERGY Z (correct)"
    ws_sum["A3"].fill = GREEN_FILL
    ws_sum["A4"] = "Red = OpenClaw Z differs from INNERGY Z (mismatch)"
    ws_sum["A4"].fill = RED_FILL
    ws_sum["A5"] = "Grey = No OpenClaw Z data available"
    ws_sum["A5"].fill = GREY_FILL
    ws_sum["A7"] = "Dimension Reference:"
    ws_sum["A8"] = "X = width  (face width)"
    ws_sum["A9"] = "Y = length (piece length)"
    ws_sum["A10"] = "Z = depth  (front-to-back)"
    ws_sum["A11"] = "Z=12 = floating shelf depth  (correct)"
    ws_sum["A12"] = "Z=25 = countertop depth  (mislabeled)"

    wb.save(output_path)
    print(f"Saved: {output_path}")


if __name__ == "__main__":
    parser = argparse.ArgumentParser()
    parser.add_argument("--xlsx", required=True)
    parser.add_argument("--output", required=True)
    args = parser.parse_args()
    main(args.xlsx, args.output)
