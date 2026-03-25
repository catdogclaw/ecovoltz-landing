#!/usr/bin/env python3
"""
completeness_check.py
Compares drawing cabinet counts against INNERGY line item counts.
Identifies: mislabeled items, missing cabinets, extra items, scope LF discrepancies.
"""

import argparse
import fitz
import openpyxl
import re
from collections import defaultdict

CABINET_MARKERS = {
    "PL1": "Plastic laminate cabinet type 1",
    "PL2": "Plastic laminate cabinet type 2",
    "SS1": "Solid surface shelf/counter type 1",
    "SS2": "Solid surface shelf/counter type 2",
    "F6":  "Floating shelf marker (F6 on drawings — verify Z depth!)",
    "F1":  "Filler/panel F1",
    "F2":  "Filler/panel F2",
    "F3":  "Filler/panel F3",
    "F4":  "Filler/panel F4",
    "F5":  "Filler/panel F5",
    # NOTE: These are MATERIAL types, NOT cabinet types:
    "P1":  "Paint finish on wall panels — NOT a cabinet type",
    "ST3": "Stone/tile backsplash material — NOT a cabinet type",
    "B1":  "Rubber base material (Johnsonite) — NOT a cabinet variant",
    "B2":  "Rubber base material (Johnsonite) — NOT a cabinet variant",
}


def get_drawing_markers(doc, page_idx):
    page = doc[page_idx]
    blocks = page.get_text("dict", flags=11)["blocks"]
    items = []
    for block in blocks:
        if "lines" not in block:
            continue
        for line in block["lines"]:
            for span in line["spans"]:
                text = span["text"].strip()
                x, y = span["origin"]
                fs = span.get("size", 0)
                if not text:
                    continue
                if text in CABINET_MARKERS:
                    items.append({"type": text, "label": CABINET_MARKERS[text], "x": x, "y": y, "fs": fs})
                if re.match(r"^\d+'-?\d*\"?$", text) or re.match(r"^\d+/\d+\"?$", text):
                    items.append({"type": "DIM", "label": text, "x": x, "y": y, "fs": fs})
    return items


def estimate_cabinet_count_by_type(doc, page_idx):
    items = get_drawing_markers(doc, page_idx)
    counts = defaultdict(int)
    for item in items:
        if item["type"] in CABINET_MARKERS:
            counts[item["type"]] += 1
    return dict(counts)


def count_innergy_items(qc_xlsx_path):
    wb = openpyxl.load_workbook(qc_xlsx_path, data_only=True)
    ws = wb["QC Comparison"]
    suite_items = defaultdict(lambda: defaultdict(int))
    for row in ws.iter_rows(min_row=2, values_only=True):
        name = str(row[1] or "")
        loc = str(row[2] or "")
        suite = loc.split(" - (")[0].strip() if " - (" in loc else loc.split(" - ")[0].strip()
        n = name.lower()
        if "tall" in n:
            itype = "TALL_CAB"
        elif "wall" in n:
            itype = "WALL_CAB"
        elif "trash" in n:
            itype = "TRASH_CAB"
        elif "base door" in n or "base drawer" in n or "base mw" in n:
            itype = "BASE_CAB"
        elif "floating shelf" in n:
            itype = "FLOATING_SHELF"
        elif "diewall" in n:
            itype = "DIEWALL"
        elif "panel filler" in n:
            itype = "PANEL_FILLER"
        elif "coat rod" in n:
            itype = "COAT_ROD"
        elif "plywood subtop" in n:
            itype = "PLYWOOD_SUBTOP"
        elif "solid surface" in n:
            itype = "SOLID_SURFACE"
        elif "add." in n or "general conditions" in n:
            itype = "IGNORE"
        elif any(k in n for k in ["top", "counter"]):
            itype = "TOP_COUNTER"
        else:
            itype = "OTHER"
        if itype != "IGNORE":
            suite_items[suite][itype] += 1
    wb.close()
    return dict(suite_items)


def check_mislabeled_floating_shelves(qc_xlsx_path):
    """
    Check floating shelf items for mislabeling.

    DIMENSION AXES (INNERGY convention in this dataset):
      X = width  (face width)
      Y = length (piece length — THIS IS THE CHECK DIMENSION in this dataset)
      Z = depth  (front-to-back, typically 2" for shelves)

    RULE: Floating shelf Y-length = 12"
           Countertop Y-length = 25" (countertop depth mislabeled as shelf)

    A "Floating Shelf PL" item with Y=25" is mislabeled — it is a countertop.
    Z=2 is the shelf thickness and is the same for both shelf and countertop.
    """
    wb = openpyxl.load_workbook(qc_xlsx_path, data_only=True)
    ws = wb["QC Comparison"]
    mislabeled = []
    correct = []
    for row in ws.iter_rows(min_row=2, values_only=True):
        name = str(row[1] or "")
        x = row[4]
        y = row[5]
        z = row[6]
        if "floating shelf" in name.lower() and y is not None:
            try:
                y_val = float(y)
                z_val = float(z) if z is not None else None
                item = {
                    "name": name,
                    "x": float(x) if x is not None else None,
                    "y": y_val,
                    "z": z_val,
                    "loc": str(row[2] or "")[:60],
                }
                if y_val == 25:
                    # Y=25 is countertop depth — mislabeled
                    item["likely"] = "COUNTERTOP (Y=25)"
                    item["signal"] = "Y=25 confirms countertop depth"
                    mislabeled.append(item)
                elif y_val == 12:
                    # Y=12 is floating shelf depth — correctly labeled
                    item["likely"] = "FLOATING SHELF (Y=12)"
                    item["signal"] = "Y=12 matches shelf depth"
                    correct.append(item)
                else:
                    # Non-standard Y length
                    item["likely"] = f"UNUSUAL (Y={y_val})"
                    item["signal"] = "Check manually"
                    mislabeled.append(item)
            except (ValueError, TypeError):
                pass
    wb.close()
    return mislabeled, correct


def run_completeness_check(drawing_pdf_path, qc_xlsx_path, drawing_page_map, output_path=None):
    doc = fitz.open(drawing_pdf_path)
    innergy_counts = count_innergy_items(qc_xlsx_path)
    mislabeled, correct = check_mislabeled_floating_shelves(qc_xlsx_path)
    lines = [
        "# Completeness Check Report",
        f"*Drawing:* `{drawing_pdf_path}`",
        f"*QC Spreadsheet:* `{qc_xlsx_path}`",
        "",
        "## IMPORTANT: Dimension Reference",
        "",
        "Before reviewing findings, confirm the INNERGY dimension axes:",
        "- **X** = width  (horizontal face width)",
        "- **Y** = length (horizontal piece length — THIS IS THE CHECK DIMENSION in this dataset)",
        "- **Z** = depth  (front-to-back, typically 2\" for floating shelves in this dataset)",
        "",
        "**Floating shelf Y-length = 12\"**",
        "**Countertop Y-length = 25\"**",
        "",
        "A mislabeled floating shelf has Y=25 (countertop length, not shelf).",
        "",
    ]

    if mislabeled:
        mis_y25 = [m for m in mislabeled if "COUNTERTOP" in m.get("likely", "")]
        lines += [
            f"## 🔴 MISLABELED Floating Shelves ({len(mis_y25)} items with Y=25 — countertops)",
            "",
            "| Suite | Name | X | Y | Z | Signal |",
            "|-------|------|---|---|---|--------|",
        ]
        for m in mislabeled:
            likely = m.get("likely", "UNKNOWN")
            signal = m.get("signal", "")
            x_val = m.get('x')
            x_str = f"{x_val:.2f}" if isinstance(x_val, float) else str(x_val or 'N/A')
            lines.append(f"| {m['loc'][:25]} | {m['name'][:40]} | {x_str} | {m.get('y', 0):.2f} | {m.get('z', 0):.2f} | {signal} |")
        lines.append("")
        lines.append(f"**🔴 Action:** Reclassify {len(mis_y25)} items as PL Tops in InnerGy. Y=25 = countertop length, not shelf.\n")

    if correct:
        lines += [
            f"## ✅ Correctly Labeled Floating Shelves ({len(correct)} items with Y=12)",
            "",
            "| Suite | Name | X | Y | Z |",
            "|-------|------|---|---|---|",
        ]
        for c in correct:
            x_v = c.get('x')
            x_str = f"{x_v:.2f}" if isinstance(x_v, float) else str(x_v or 'N/A')
            lines.append(f"| {c['loc'][:25]} | {c['name'][:40]} | {x_str} | {c.get('y', 0):.2f} | {c.get('z', 0):.2f} |")
        lines.append("")
        lines.append(f"**✅ These {len(correct)} items are correctly labeled.** Y=12 matches floating shelf length.\n")

    lines += [
        "## Drawing Cabinet Marker Counts (from A401 elevation sheets)",
        "",
        "| Marker | Count | Description |",
        "|--------|-------|-------------|",
    ]
    for suite_name, page_idx in sorted(drawing_page_map.items()):
        counts = estimate_cabinet_count_by_type(doc, page_idx)
        lines.append(f"### {suite_name} (source page {page_idx + 1})")
        lines.append("")
        if counts:
            for marker, count in sorted(counts.items(), key=lambda x: -x[1]):
                lines.append(f"| `{marker}` | {count} | {CABINET_MARKERS.get(marker, 'N/A')} |")
        else:
            lines.append("*No cabinet type markers found on this page.*")
        lines.append("")

    lines += [
        "## INNERGY Line Item Counts",
        "",
        "| Suite | Item Type | Count |",
        "|-------|-----------|-------|",
    ]
    for suite, items in sorted(innergy_counts.items()):
        for itype, count in sorted(items.items(), key=lambda x: -x[1]):
            lines.append(f"| {suite[:35]} | {itype} | {count} |")

    lines += [
        "",
        "## Completeness Assessment",
        "",
        "### Questions to Answer per Suite:",
        "1. **Base cabinets:** Count P1 markers on drawing vs. `BASE_CAB + TRASH_CAB` in INNERGY per suite",
        "2. **Wall cabinets:** Count ST3 markers on drawing vs. `WALL_CAB` in INNERGY per suite",
        "3. **Tall cabinets:** Confirm 90\" tall cabinets exist in drawings (not just in scope spec)",
        "4. **Floating shelves:** Confirm actual shelf dimensions match INNERGY floating shelf line items",
        "5. **Tops/Countertops:** Counters should appear as PL Top items, NOT floating shelves",
        "6. **Scope LF:** Compare stated LF in scope vs. sum of cabinet face widths in INNERGY",
        "",
        "## Common Patterns",
        "",
        "| Pattern | Likely Cause | Action |",
        "|---------|-------------|--------|",
        "| Drawing has base cabs INNERGY doesn't | Cabinets missed in takeoff | Add to InnerGy |",
        "| INNERGY has items not in drawing | Scope item not shown in drawings | Verify with GC/architect |",
        "| Mislabeled floating shelf Y=25 | Countertop length (Y=25) mislabeled as shelf (Y=12) | Reclassify as PL Top |",
        "| Scope LF >> actual cab widths | Budgeted vs. drawn footage | Note variance in bid |",
    ]

    doc.close()
    report = "\n".join(lines)
    if output_path:
        with open(output_path, "w") as f:
            f.write(report)
        print(f"Completeness report saved: {output_path}")
    else:
        print(report)
    return report


if __name__ == "__main__":
    parser = argparse.ArgumentParser()
    parser.add_argument("--drawing", required=True)
    parser.add_argument("--innergy", required=True)
    parser.add_argument("--output", help="Path for output markdown report")
    args = parser.parse_args()
    # Page indices for drawings_extracted.pdf (4 pages, 0-indexed):
    # Page 0 = Suite 1200, Page 1 = Suite 1150, Page 2 = Suite 1275/1285, Page 3 = Suite 1300
    page_map = {
        "Suite 1200": 0,
        "Suite 1150": 1,
        "Suite 1275/1285": 2,
        "Suite 1300": 3,
    }
    run_completeness_check(args.drawing, args.innergy, page_map, args.output)
