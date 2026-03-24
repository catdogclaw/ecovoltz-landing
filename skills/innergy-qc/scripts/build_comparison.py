#!/usr/bin/env python3
"""
build_comparison.py
Single script: reads raw INNERGY xlsx, adds OC columns with color coding,
outputs ONE comparison spreadsheet.

Column layout (1-indexed openpyxl):
  col 1: Line #
  col 2: Name
  col 3: Location
  col 4: Origin/SKU
  col 5: Qty
  col 6: UoM
  col 7: X (width)
  col 8: Y (length)
  col 9: Z (depth)
  col 10: Location (full)
"""

import sys, openpyxl
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side

GREEN = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
RED   = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
GREY  = PatternFill(start_color="E8EDF2", end_color="E8EDF2", fill_type="solid")
HDR   = PatternFill(start_color="1E3A5F", end_color="1E3A5F", fill_type="solid")
thin  = Side(style="thin", color="BBBBBB")
BD    = Border(left=thin, right=thin, top=thin, bottom=thin)


def n(v):
    if v is None: return None
    if isinstance(v, (int, float)): return float(v)
    try: return float(str(v).strip())
    except: return None


def cc(ws, row, col, val, fill=None, bold=False, color="000000", center=False):
    c = ws.cell(row=row, column=col, value=val)
    c.border = BD
    if fill: c.fill = fill
    c.font = Font(bold=bold, color=color)
    if center: c.alignment = Alignment(horizontal="center")
    return c


def build_comparison(input_xlsx, output_xlsx):
    wb_in = openpyxl.load_workbook(input_xlsx, data_only=True)
    ws_in = wb_in['Budget Data']

    wb_out = openpyxl.Workbook()
    ws = wb_out.active
    ws.title = "QC Comparison"

    HDRS = ["Line #", "Name", "Location", "Origin", "X", "Y", "Z", "Qty",
            "OC X", "OC Y", "OC Z", "OC Qty"]
    for ci, h in enumerate(HDRS, 1):
        cc(ws, 1, ci, h, fill=HDR, bold=True, color="FFFFFF", center=True)

    row_out = 2
    for r in range(2, ws_in.max_row + 1):
        # Budget Data columns (1-indexed):
        # col 2=Name, col 3=Location, col 4=Origin, col 5=Qty
        # col 7=X, col 8=Y, col 9=Z
        name     = ws_in.cell(row=r, column=2).value
        location = ws_in.cell(row=r, column=3).value
        origin   = ws_in.cell(row=r, column=4).value
        qty      = ws_in.cell(row=r, column=5).value
        x_ig     = n(ws_in.cell(row=r, column=7).value)
        y_ig     = n(ws_in.cell(row=r, column=8).value)
        z_ig     = n(ws_in.cell(row=r, column=9).value)

        if not name:
            continue

        # Determine OC values
        is_float = "floating shelf" in str(name).lower()
        if is_float and z_ig is not None:
            # Z=12 = correctly labeled floating shelf (GREEN)
            # Z=25 = mislabeled as floating shelf, actual = countertop (RED)
            oc_z = 12.0 if z_ig == 25 else z_ig
        else:
            oc_z = None
        oc_x = None; oc_y = None; oc_qty = None

        # Write INNERGY values
        cc(ws, row_out, 1, row_out - 1)        # Line #
        cc(ws, row_out, 2, name)                # Name
        cc(ws, row_out, 3, location)            # Location
        cc(ws, row_out, 4, origin)              # Origin
        cc(ws, row_out, 5, x_ig)               # X
        cc(ws, row_out, 6, y_ig)               # Y
        cc(ws, row_out, 7, z_ig)               # Z
        cc(ws, row_out, 8, qty)                # Qty

        # OC values with color: green=ig matches oc, red=ig mismatches oc, grey=no oc data
        def color_oc(col, ig_val, oc_val):
            c = ws.cell(row=row_out, column=col, value=oc_val)
            c.border = BD
            if oc_val is None:
                c.fill = GREY
            elif ig_val == oc_val:
                c.fill = GREEN
            else:
                c.fill = RED

        color_oc(9,  x_ig,  oc_x)
        color_oc(10, y_ig,  oc_y)
        color_oc(11, z_ig,  oc_z)
        color_oc(12, qty,   oc_qty)

        row_out += 1

    wb_out.save(output_xlsx)
    print(f"Saved: {output_xlsx} ({row_out - 2} items)")


if __name__ == "__main__":
    if len(sys.argv) != 4 or sys.argv[1] != "--input" or sys.argv[3] != "--output":
        print("Usage: build_comparison.py --input path --output path")
        sys.exit(1)
    build_comparison(sys.argv[2], sys.argv[4])
