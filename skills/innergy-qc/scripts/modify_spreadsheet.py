#!/usr/bin/env python3
"""
modify_spreadsheet.py
Reads INNERGY estimating xlsx → extracts all line items with X/Y/Z/qty/suite
→ outputs clean xlsx for QC comparison.
"""

import argparse
import openpyxl
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from openpyxl.utils import get_column_letter

HEADER_FILL = PatternFill(start_color="1E3A5F", end_color="1E3A5F", fill_type="solid")
GREY_FILL   = PatternFill(start_color="E8EDF2", end_color="E8EDF2", fill_type="solid")
thin = Side(style="thin", color="BBBBBB")
border = Border(left=thin, right=thin, top=thin, bottom=thin)


def extract_innergy(input_path):
    """Read INNERGY xlsx, extract line items."""
    wb = openpyxl.load_workbook(input_path, data_only=True)

    sheets_to_try = ["Budget Data", "Budget", "Line Items", "Items"]
    ws = None
    for sname in sheets_to_try:
        if sname in wb.sheetnames:
            ws = wb[sname]
            break
    if ws is None:
        raise ValueError(f"Could not find budget data sheet. Available: {wb.sheetnames}")

    items = []
    for row in ws.iter_rows(min_row=2, values_only=True):
        if not row[0] and not row[1]:
            continue
        # Budget Data ACTUAL column layout (verified against raw data):
        # col[0]=Origin Name, col[1]=Name, col[2]=Library Name, col[3]=SKU,
        # col[4]=E (Qty), col[5]=F (X), col[6]=G (Y), col[7]=H (Z), col[8]=I (qty-count), col[9]=J (Location)
        name = str(row[1] or "").strip()   # col B = Name
        if not name:
            continue
        location = str(row[9] or "").strip()  # col J = Location
        origin = str(row[3] or "").strip()   # col D = SKU
        qty = row[4] if len(row) > 4 else 1  # col E = Qty
        unit = str(row[5] or "").strip() if len(row) > 5 else "EA"  # col F = X (used as unit when X is empty)
        x = row[5] if len(row) > 5 else None  # col F = X dimension
        y = row[6] if len(row) > 6 else None  # col G = Y dimension
        z = row[7] if len(row) > 7 else None  # col H = Z dimension
        price = row[19] if len(row) > 19 else None  # col T = Extended Labor Cost (or use col[16]?)
        items.append({
            "name": name,
            "location": location,
            "origin": origin,
            "qty": qty or 1,
            "unit": unit,
            "x": x,
            "y": y,
            "z": z,
            "price": price,
        })

    wb.close()
    return items


def create_qc_spreadsheet(items, output_path):
    """Create clean QC xlsx with headers for OC columns."""
    wb = openpyxl.Workbook()

    # ── QC Comparison sheet ──────────────────────────────────────────────
    ws = wb.active
    ws.title = "QC Comparison"

    headers = [
        "Line #", "Name", "Location", "Origin",
        "X (in)", "Y (in)", "Z (in)", "Qty",
        "OC X", "OC Y", "OC Z", "OC Qty",
        "X Var", "Y Var", "Z Var",
        "Status", "Notes"
    ]

    for col, hdr in enumerate(headers, 1):
        c = ws.cell(row=1, column=col, value=hdr)
        c.fill = HEADER_FILL
        c.font = Font(bold=True, color="FFFFFF", size=11)
        c.alignment = Alignment(horizontal="center", wrap_text=True)
        c.border = border

    for i, item in enumerate(items, 1):
        for col, val in enumerate([
            i,
            item["name"],
            item["location"],
            item["origin"],
            item["x"],
            item["y"],
            item["z"],
            item["qty"],
        ], 1):
            c = ws.cell(row=i+1, column=col, value=val)
            c.border = border
            if col == 1:
                c.alignment = Alignment(horizontal="center")
            if col >= 5:
                c.fill = GREY_FILL

    # ── QC Summary sheet ─────────────────────────────────────────────────
    ws_sum = wb.create_sheet("QC Summary")
    ws_sum["A1"] = "QC Summary"
    ws_sum["A1"].font = Font(bold=True, size=14)

    # Column widths
    ws.column_dimensions["A"].width = 6
    ws.column_dimensions["B"].width = 45
    ws.column_dimensions["C"].width = 45
    ws.column_dimensions["D"].width = 12
    ws.column_dimensions["E"].width = 12
    ws.column_dimensions["F"].width = 12
    ws.column_dimensions["G"].width = 12
    ws.column_dimensions["H"].width = 8
    ws.column_dimensions["I"].width = 12
    ws.column_dimensions["J"].width = 12
    ws.column_dimensions["K"].width = 12
    ws.column_dimensions["L"].width = 12
    ws.column_dimensions["M"].width = 12
    ws.column_dimensions["N"].width = 12
    ws.column_dimensions["O"].width = 12
    ws.column_dimensions["P"].width = 16
    ws.column_dimensions["Q"].width = 45

    ws_sum.column_dimensions["A"].width = 40
    ws_sum.column_dimensions["B"].width = 14
    ws_sum.column_dimensions["C"].width = 18

    wb.save(output_path)
    print(f"Saved: {output_path}  ({len(items)} line items)")


if __name__ == "__main__":
    parser = argparse.ArgumentParser(description="Extract INNERGY line items to QC xlsx")
    parser.add_argument("--input", required=True, help="Path to INNERGY xlsx")
    parser.add_argument("--output", required=True, help="Path for output xlsx")
    args = parser.parse_args()

    items = extract_innergy(args.input)
    create_qc_spreadsheet(items, args.output)
