#!/usr/bin/env python3
"""
populate_qc_xlsx.py
Populates the innergy_qc.xlsx with OpenClaw dimension findings.
Reads from scratch/findings_*.md per-suite notes, or from a structured
FINDINGS dict defined in the script.

Usage:
    python3 scripts/populate_qc_xlsx.py \
        --xlsx path/to/innergy_qc.xlsx \
        --output path/to/output.xlsx
"""

import argparse
import openpyxl
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side

GREEN_FILL  = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
RED_FILL    = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
ORANGE_FILL = PatternFill(start_color="FFEB9C", end_color="FFEB9C", fill_type="solid")
GREY_FILL   = PatternFill(start_color="E8EDF2", end_color="E8EDF2", fill_type="solid")
DIM_FILL    = PatternFill(start_color="D9E1F2", end_color="D9E1F2", fill_type="solid")
HEADER_FILL = PatternFill(start_color="1E3A5F", end_color="1E3A5F", fill_type="solid")
thin = Side(style="thin", color="BBBBBB")
border = Border(left=thin, right=thin, top=thin, bottom=thin)


def num(v):
    if v is None:
        return None
    if isinstance(v, (int, float)):
        return float(v)
    s = str(v).strip().lstrip("~")
    if s == "" or s in ("N/A", "None"):
        return None
    try:
        return float(s)
    except ValueError:
        return None


def variance_str(ig_val, oc_val):
    ig_n = num(ig_val)
    oc_n = num(oc_val)
    if ig_n is None or oc_n is None:
        return "—"
    return f"{num(oc_val) - ig_n:+.2f}\""


def status_style(cell, status):
    cell.border = border
    cell.alignment = Alignment(horizontal="center")
    if status == "CONFIRMED":
        cell.fill = GREEN_FILL
        cell.font = Font(bold=True, color="276221")
    elif status == "DISCREPANCY":
        cell.fill = RED_FILL
        cell.font = Font(bold=True, color="9C0006")
    elif status in ("MISSING_IN_DRAWING", "MISSING_IN_INNERGY"):
        cell.fill = ORANGE_FILL
        cell.font = Font(bold=True, color="843C0C")
    elif status == "NEEDS REVIEW":
        cell.fill = ORANGE_FILL
        cell.font = Font(bold=True, color="843C0C")
    else:
        cell.fill = GREY_FILL
        cell.font = Font(size=9)


def variance_color(cell, val, threshold=0.5):
    cell.border = border
    if val is None or val == "" or val == "—" or "?" in str(val):
        return
    try:
        s = str(val).rstrip('"').strip()
        v = abs(float(s))
        if v > threshold:
            cell.fill = RED_FILL
            cell.font = Font(color="9C0006")
        else:
            cell.fill = GREEN_FILL
            cell.font = Font(color="276221")
    except ValueError:
        pass


def fill_cell(ws, row, col, val, fill=None, font_kw=None):
    c = ws.cell(row=row, column=col, value=val)
    c.border = border
    if fill:
        c.fill = fill
    if font_kw:
        c.font = Font(**font_kw)
    return c


def match_findings(name, location, x_ig, y_ig, z_ig, qty_ig, findings):
    """
    Match a spreadsheet row to the FINDINGS table.
    Returns dict with keys: oc_x, oc_y, oc_z, oc_qty, status, notes
    """
    n_lower = name.lower()
    loc_upper = location.upper()
    x_ig_n = num(x_ig)
    z_ig_n = num(z_ig)

    candidates = []
    for f in findings:
        np = f.get("name_pat", "").lower()
        sp = f.get("suite_pat", "").upper()
        if np not in n_lower and n_lower not in np:
            continue
        if sp and sp not in loc_upper:
            continue
        if x_ig_n is not None and f.get('ig_x_min', 0) > 0:
            if not (f['ig_x_min'] - 0.5 <= x_ig_n <= f['ig_x_max'] + 0.5):
                continue
        candidates.append(f)

    if not candidates:
        return None
    return max(candidates, key=lambda f: (len(f.get("name_pat", "")), len(f.get("suite_pat", ""))))


def populate(xlsx_path, output_path, findings):
    wb = openpyxl.load_workbook(xlsx_path)
    ws = wb["QC Comparison"]

    suite_stats = {}

    for row_idx in range(2, ws.max_row + 1):
        name = ws.cell(row=row_idx, column=2).value or ""
        location = ws.cell(row=row_idx, column=3).value or ""
        x_ig = ws.cell(row=row_idx, column=5).value
        y_ig = ws.cell(row=row_idx, column=6).value
        z_ig = ws.cell(row=row_idx, column=7).value
        qty_ig = ws.cell(row=row_idx, column=8).value

        finding = match_findings(name, location, x_ig, y_ig, z_ig, qty_ig, findings)
        sk = location.split(" - (")[0].strip() if location else "Unknown"
        if suite_stats.get(sk) is None:
            suite_stats[sk] = {"confirmed": 0, "discrepancy": 0, "needs_review": 0, "na": 0}

        if finding:
            oc_x = finding.get("x")
            oc_y = finding.get("y")
            oc_z = finding.get("z")
            oc_qty = finding.get("qty")
            status = finding.get("status", "NEEDS REVIEW")
            notes = finding.get("notes", "")

            fill_cell(ws, row_idx, 9,  oc_x, fill=DIM_FILL)
            fill_cell(ws, row_idx, 10, oc_y, fill=DIM_FILL)
            fill_cell(ws, row_idx, 11, oc_z, fill=DIM_FILL)
            fill_cell(ws, row_idx, 12, oc_qty if oc_qty != "N/A" else None, fill=DIM_FILL)

            x_var = variance_str(x_ig, oc_x)
            y_var = variance_str(y_ig, oc_y)
            z_var = variance_str(z_ig, oc_z)
            fill_cell(ws, row_idx, 13, x_var)
            fill_cell(ws, row_idx, 14, y_var)
            fill_cell(ws, row_idx, 15, z_var)
            variance_color(ws.cell(row=row_idx, column=13), x_var)
            variance_color(ws.cell(row=row_idx, column=14), y_var)
            variance_color(ws.cell(row=row_idx, column=15), z_var)

            status_cell = ws.cell(row=row_idx, column=16, value=status)
            status_style(status_cell, status)

            notes_cell = ws.cell(row=row_idx, column=17, value=notes)
            notes_cell.font = Font(size=9)
            notes_cell.border = border

            if status == "CONFIRMED":
                suite_stats[sk]["confirmed"] += 1
            elif status == "DISCREPANCY":
                suite_stats[sk]["discrepancy"] += 1
            elif status in ("NEEDS REVIEW", "MISSING_IN_DRAWING", "MISSING_IN_INNERGY"):
                suite_stats[sk]["needs_review"] += 1
            elif status == "N/A_PRICING":
                suite_stats[sk]["na"] += 1
        else:
            status_cell = ws.cell(row=row_idx, column=16, value="NEEDS REVIEW")
            status_style(status_cell, "NEEDS REVIEW")
            notes_cell = ws.cell(row=row_idx, column=17, value="No matching findings — review manually")
            notes_cell.font = Font(size=9)
            notes_cell.border = border
            suite_stats[sk]["needs_review"] += 1

    # ── Fill QC Summary tab ───────────────────────────────────────────────
    ws_sum = wb["QC Summary"]
    ws_sum["A1"] = "QC Summary — Generated by OpenClaw innergy-qc skill"
    ws_sum["A1"].font = Font(bold=True, size=14)
    ws_sum["A2"] = "Generated: 2026-03-24"
    ws_sum["A4"] = "Suite / Area"
    ws_sum["B4"] = "Confirmed"
    ws_sum["C4"] = "Discrepancies"
    ws_sum["D4"] = "Needs Review"
    ws_sum["E4"] = "N/A"
    ws_sum["F4"] = "Key Issue"

    for c in [ws_sum["A4"], ws_sum["B4"], ws_sum["C4"], ws_sum["D4"], ws_sum["E4"], ws_sum["F4"]]:
        c.fill = HEADER_FILL
        c.font = Font(bold=True, color="FFFFFF")
        c.alignment = Alignment(horizontal="center")
        c.border = border

    ws_sum.column_dimensions["F"].width = 50
    row = 5
    for suite, stats in sorted(suite_stats.items()):
        ws_sum.cell(row=row, column=1, value=suite).border = border
        for col, key in [(2, "confirmed"), (3, "discrepancy"), (4, "needs_review"), (5, "na")]:
            c = ws_sum.cell(row=row, column=col, value=stats[key])
            c.border = border
            if stats[key] == 0:
                c.fill = GREEN_FILL
                c.font = Font(color="276221")
            elif key == "discrepancy" and stats[key] > 0:
                c.fill = RED_FILL
                c.font = Font(color="9C0006")
            elif key in ("needs_review",) and stats[key] > 0:
                c.fill = ORANGE_FILL
                c.font = Font(color="843C0C")
        row += 1

    wb.save(output_path)
    print(f"Saved: {output_path}")
    print("\nSuite Summary:")
    for suite, stats in sorted(suite_stats.items()):
        total = sum(stats.values())
        print(f"  {suite}: {stats['confirmed']} confirmed, {stats['discrepancy']} discrepancies, {stats['needs_review']} needs review / {total} total")


if __name__ == "__main__":
    parser = argparse.ArgumentParser()
    parser.add_argument("--xlsx", required=True, help="Path to innergy_qc.xlsx")
    parser.add_argument("--output", required=True, help="Path for output xlsx")
    args = parser.parse_args()

    # FINDINGS — 595 Market Street Run 3
    # Dimension axes: X=width, Y=length, Z=depth (front-to-back)
    # Floating shelf rule: Z=12 = shelf (CORRECT), Z=25 = countertop (MISLABELED)
    FINDINGS = [
        # === CORRECTLY LABELED FLOATING SHELVES (Z=12) ===
        {
            "name_pat": "floating shelf pl",
            "ig_x_min": 4.0, "ig_x_max": 5.0,
            "z": 12,
            "status": "CONFIRMED",
            "notes": "Z=12 — correct floating shelf depth. ✅ CORRECT."
        },
        {
            "name_pat": "floating shelf pl",
            "ig_x_min": 6.0, "ig_x_max": 6.5,
            "z": 12,
            "status": "CONFIRMED",
            "notes": "Z=12 — correct floating shelf depth. ✅ CORRECT."
        },
        {
            "name_pat": "floating shelf pl",
            "ig_x_min": 7.0, "ig_x_max": 8.5,
            "z": 12,
            "status": "CONFIRMED",
            "notes": "Z=12 — correct floating shelf depth. ✅ CORRECT."
        },
        # === MISLABELED FLOATING SHELVES (Z=25 = countertop depth) ===
        {
            "name_pat": "floating shelf pl",
            "ig_x_min": 6.0, "ig_x_max": 6.5,
            "z": 25,
            "status": "DISCREPANCY",
            "notes": "Z=25 — countertop depth, not shelf. 🔴 MISLABELED. Reclassify as PL Top."
        },
        {
            "name_pat": "floating shelf pl",
            "ig_x_min": 7.5, "ig_x_max": 8.0,
            "z": 25,
            "status": "DISCREPANCY",
            "notes": "Z=25 — countertop depth, not shelf. 🔴 MISLABELED. Reclassify as PL Top."
        },
        # === BASE CABINETS — confirmed heights ===
        {
            "name_pat": "pl base door",
            "status": "CONFIRMED",
            "notes": "Base cabinet body height ~32.5\" + 1.5\" top = 34\" standard. ✅ OK."
        },
        {
            "name_pat": "pl trash cabinet",
            "status": "CONFIRMED",
            "notes": "T/R/C trash cabinet — standard base height. ✅ OK."
        },
        {
            "name_pat": "pl tall door",
            "status": "CONFIRMED",
            "notes": "Tall cabinet — confirmed in drawings. ✅ OK."
        },
        {
            "name_pat": "diewall",
            "status": "CONFIRMED",
            "notes": "DieWall pieces confirmed — island framing. ✅ OK."
        },
        {
            "name_pat": "coat rod",
            "status": "CONFIRMED",
            "notes": "Coat rod confirmed in Suite 1275 SW. ✅ OK."
        },
        {
            "name_pat": "panel filler",
            "status": "CONFIRMED",
            "notes": "Panel filler confirmed in drawings. ✅ OK."
        },
        {
            "name_pat": "plywood subtop",
            "status": "CONFIRMED",
            "notes": "Plywood subtop — standard for solid surface counters. ✅ OK."
        },
        {
            "name_pat": "solid surface",
            "status": "CONFIRMED",
            "notes": "Solid surface top (SS1/SS2). ✅ OK."
        },
        {
            "name_pat": "pl wall door cabinet",
            "status": "CONFIRMED",
            "notes": "Wall cabinet — confirmed present. ✅ OK."
        },
        {
            "name_pat": "add.",
            "status": "N/A_PRICING",
            "notes": "Allowance/additive item — not verified against drawings."
        },
    ]

    populate(args.xlsx, args.output, FINDINGS)
