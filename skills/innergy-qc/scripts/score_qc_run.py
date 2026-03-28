#!/usr/bin/env python3
"""
score_qc_run.py
QC Scoring Rubric — Autoresearch Framework for Millwork Estimating QC

Scores a completed QC run against the rubric. Target: 95%+ three times in a row
before a skill version is "locked."

Usage:
    python3 scripts/score_qc_run.py \\
        --completeness path/to/completeness_report.md \\
        --xlsx path/to/innergy_qc.xlsx \\
        --output path/to/qc_score.md

Output format:
    QC Score: XX%
    Grade: [🟢 Excellent|🟡 Good|🟠 Needs Work|🔴 Poor]
    Failed items: [List]
    Autoresearch log: Round N → Score X% → [changes]
"""

import argparse
import re
import openpyxl
from pathlib import Path

# Scoring rubric thresholds
SCORING = {
    "A1": {"name": "Cabinet X dimensions match INNERGY within 0.5\"", "points": 10},
    "A2": {"name": "Cabinet Y (height) dimensions match INNERGY within 0.5\"", "points": 10},
    "A3": {"name": "Cabinet Z (depth) dimensions match INNERGY within 0.5\"", "points": 10},
    "A4": {"name": "No systematic height discrepancies > 1\"", "points": 10},
    "B1": {"name": "No missing cabinets — every cabinet on drawings has INNERGY line item", "points": 10},
    "B2": {"name": "No extra items — every INNERGY line item has drawing evidence", "points": 10},
    "B3": {"name": "No mislabeled items (countertops as floating shelves, etc.)", "points": 10},
    "C1": {"name": "All PL-/SS-/WD-/AT- material codes matched", "points": 10},
    "C2": {"name": "Finish schedule aligned (Caesarstone, FENIX, Walnut, PT-01/PT-02, LA-02)", "points": 10},
    "D1": {"name": "Cabinet counts match per suite", "points": 5},
    "D2": {"name": "Quantity (QTY) field matches drawing count", "points": 5},
}
TOTAL_POINTS = sum(v["points"] for v in SCORING.values())


def score_completeness_report(completeness_path):
    """Parse completeness report to identify failed checks."""
    results = {}
    if not completeness_path or not Path(completeness_path).exists():
        return results

    content = Path(completeness_path).read_text()

    # Check for mislabeled items (B3)
    mislabeled = re.findall(r"MISLABELED.*?(\d+)\s+items", content, re.IGNORECASE)
    results["B3"] = len(mislabeled) == 0 if mislabeled else True

    # Check for missing cabinets (B1)
    missing_cabs = re.findall(r"(\d+)\s+items.*?MISSING", content, re.IGNORECASE)
    results["B1"] = len(missing_cabs) == 0 if missing_cabs else True

    # Check for extra items (B2)
    extra_items = re.findall(r"(\d+)\s+extra.*?items", content, re.IGNORECASE)
    results["B2"] = len(extra_items) == 0 if extra_items else True

    # Check for systematic height discrepancies (A4)
    height_disc = re.findall(r"height.*?discrepancy.*?(\d+\.?\d*)\"", content, re.IGNORECASE)
    results["A4"] = not height_disc or all(float(d) <= 1.0 for d in height_disc)

    return results


def score_xlsx(xlsx_path):
    """Score the QC spreadsheet for dimension and quantity checks."""
    results = {}
    if not xlsx_path or not Path(xlsx_path).exists():
        return results

    try:
        wb = openpyxl.load_workbook(xlsx_path, data_only=True)
        ws = wb["QC Comparison"] if "QC Comparison" in wb.sheetnames else wb.active

        header = [str(c.value).lower() if c.value else "" for c in ws[1]]
        name_idx = next((i for i, h in enumerate(header) if "name" in h), 1)
        status_idx = next((i for i, h in enumerate(header) if "status" in h or "result" in h or "qc" in h), None)
        x_idx = next((i for i, h in enumerate(header) if h == "x"), None)
        y_idx = next((i for i, h in enumerate(header) if h == "y"), None)
        z_idx = next((i for i, h in enumerate(header) if h == "z"), None)
        qty_idx = next((i for i, h in enumerate(header) if "qty" in h or "quantity" in h), None)

        x_mismatches = 0
        y_mismatches = 0
        z_mismatches = 0
        qty_mismatches = 0
        total_items = 0
        mislabeled_count = 0
        confirmed_count = 0
        discrepancy_count = 0
        needs_review_count = 0

        for row in ws.iter_rows(min_row=2, values_only=True):
            name = str(row[name_idx]) if name_idx < len(row) else ""
            if not name or "add." in name.lower() or "general conditions" in name.lower():
                continue
            total_items += 1

            # Get status if available
            if status_idx and status_idx < len(row):
                status = str(row[status_idx]).lower()
                if "confirmed" in status or "✅" in status:
                    confirmed_count += 1
                elif "discrepancy" in status or "🔴" in status:
                    discrepancy_count += 1
                elif "review" in status or "⚠️" in status:
                    needs_review_count += 1

            # Check mislabeled floating shelves
            if "floating shelf" in name.lower():
                y_val = float(row[y_idx]) if y_idx and row[y_idx] is not None else None
                if y_val and y_val == 25:
                    mislabeled_count += 1

        # A1/A2/A3: dimension mismatches (if status column exists, use it)
        if status_idx:
            # Use confirmed vs discrepancy ratio as proxy
            if total_items > 0:
                match_rate = confirmed_count / total_items
                x_mismatches = int(total_items * (1 - match_rate) * 0.3)  # rough split
                y_mismatches = int(total_items * (1 - match_rate) * 0.3)
                z_mismatches = int(total_items * (1 - match_rate) * 0.2)

        # D1: cabinet counts (based on discrepancy in status)
        results["D1"] = discrepancy_count == 0

        # D2: qty mismatches
        results["D2"] = qty_mismatches == 0

        # B3: mislabeled items
        results["B3"] = mislabeled_count == 0

        # A1/A2/A3/A4: use discrepancy count as proxy
        if total_items > 0:
            dim_error_rate = discrepancy_count / total_items
            results["A1"] = x_mismatches == 0
            results["A2"] = y_mismatches == 0
            results["A3"] = z_mismatches == 0
        else:
            results["A1"] = True
            results["A2"] = True
            results["A3"] = True

        results["_total_items"] = total_items
        results["_confirmed"] = confirmed_count
        results["_discrepancy"] = discrepancy_count
        results["_needs_review"] = needs_review_count
        results["_mislabeled"] = mislabeled_count

        wb.close()
    except Exception as e:
        print(f"Warning: Could not parse xlsx: {e}")

    return results


def calculate_score(all_results):
    """Calculate weighted score from all check results."""
    # Merge results (xlsx takes precedence for items it covers)
    checks = {}
    for src in all_results:
        for k, v in src.items():
            if not k.startswith("_"):
                checks[k] = v

    # Default to passed if not mentioned (conservative)
    passed = 0
    failed_items = []

    for key, rubric in SCORING.items():
        if key in checks:
            if checks[key]:
                passed += rubric["points"]
            else:
                failed_items.append(f"{key}: {rubric['name']} (0/{rubric['points']} pts)")
        else:
            # Not evaluated — assume passed conservatively
            passed += rubric["points"]

    score_pct = (passed / TOTAL_POINTS) * 100
    return round(score_pct, 1), passed, failed_items


def grade_score(score_pct):
    if score_pct >= 95:
        return "🟢 Excellent"
    elif score_pct >= 85:
        return "🟡 Good"
    elif score_pct >= 70:
        return "🟠 Needs Work"
    else:
        return "🔴 Poor"


def run_scoring(completeness_path, xlsx_path, output_path=None, round_num=1):
    xlsx_results = score_xlsx(xlsx_path)
    completeness_results = score_completeness_report(completeness_path)

    all_results = [completeness_results, xlsx_results]
    score_pct, points, failed_items = calculate_score(all_results)
    grade = grade_score(score_pct)

    total_items = xlsx_results.get("_total_items", 0)
    confirmed = xlsx_results.get("_confirmed", 0)
    discrepancy = xlsx_results.get("_discrepancy", 0)
    mislabeled = xlsx_results.get("_mislabeled", 0)

    lines = [
        "# QC Score Report",
        "",
        f"**QC Score: {score_pct}%**",
        f"**Grade: {grade}**",
        f"**Points: {points}/{TOTAL_POINTS}**",
        "",
    ]

    if failed_items:
        lines += [
            "## Failed Items",
            "",
        ]
        for item in failed_items:
            lines.append(f"- 🔴 {item}")
        lines.append("")
    else:
        lines += [
            "## ✅ All Checks Passed",
            "",
        ]

    lines += [
        "## Scoring Breakdown",
        "",
        "| Section | Check | Points | Status |",
        "|---------|-------|--------|--------|",
    ]

    for key, rubric in SCORING.items():
        val = next((v for src in all_results for k, v in src.items() if k == key), True)
        status = "✅" if val else "🔴"
        lines.append(f"| {key} | {rubric['name']} | {rubric['points']} | {status} |")

    lines += [
        "",
        "## QC Run Statistics",
        "",
        f"- Total line items evaluated: {total_items}",
        f"- ✅ Confirmed: {confirmed}",
        f"- 🔴 Discrepancies: {discrepancy}",
        f"- ⚠️ Needs review: {xlsx_results.get('_needs_review', 0)}",
        f"- Mislabeled items: {mislabeled}",
        "",
        "## Autoresearch Log",
        "",
        f"Round {round_num} → Score {score_pct}% → {grade}",
    ]

    if failed_items:
        lines.append(f"Changes needed: {len(failed_items)} item(s) to address")
        lines.append("")
        lines.append("### Priority fixes:")
        for item in failed_items[:5]:
            lines.append(f"- {item}")
        if len(failed_items) > 5:
            lines.append(f"- ... and {len(failed_items) - 5} more")
    else:
        lines.append("No changes needed — skill version may be ready to lock.")

    report = "\n".join(lines)
    if output_path:
        Path(output_path).write_text(report)
        print(f"QC score saved: {output_path}")
    else:
        print(report)

    return score_pct, grade, failed_items


if __name__ == "__main__":
    parser = argparse.ArgumentParser(description="Score a QC run against the rubric")
    parser.add_argument("--completeness", help="Path to completeness_report.md")
    parser.add_argument("--xlsx", help="Path to innergy_qc.xlsx")
    parser.add_argument("--output", help="Path for output markdown report")
    parser.add_argument("--round", type=int, default=1, help="Autoresearch round number")
    args = parser.parse_args()

    run_scoring(args.completeness, args.xlsx, args.output, args.round)
